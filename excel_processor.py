from dataclasses import dataclass, field
from enum import Enum
from typing import Optional
import json
import logging
import requests

import openpyxl

log = logging.getLogger("excel_processor")


class RowType(Enum):
    CATEGORY = "category"
    PRODUCT = "product"
    COMBO = "combo"
    COMBO_CHILD = "combo_child"
    SKIP = "skip"


@dataclass
class ClassifiedRow:
    row_index: int
    row_type: RowType
    ma_san_pham: str = ""
    description: str = ""
    thue_vat: str = ""
    gia_ban_ra: str = ""
    gia_niem_yet: str = ""
    danh_muc: str = ""
    combo_codes: list[str] = field(default_factory=list)
    parent_combo_desc: str = ""


def find_header_row(ws) -> Optional[int]:
    """Find the header row containing 'STT' in column A.

    Iterates from row 1, returns the 1-based row number where column A
    contains 'STT' (case-insensitive). Returns None if not found.
    """
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value is not None and str(cell.value).strip().upper() == "STT":
            return cell.row
    return None


def is_bold_row(row) -> bool:
    """Check if the first non-empty cell in the row has bold font.

    Args:
        row: An openpyxl row tuple (from iter_rows).

    Returns:
        True if the first non-empty cell has bold font, False otherwise.
    """
    for cell in row:
        if cell.value is not None and str(cell.value).strip() != "":
            if cell.font and cell.font.bold:
                return True
            return False
    return False


def is_combo_row(row, col_indices: dict) -> bool:
    """Check if a row is a combo row (Nhóm hàng = 'MIX' AND Thuế VAT = 'MIX').

    Args:
        row: An openpyxl row tuple (from iter_rows).
        col_indices: Dict mapping column names to 0-based indices.

    Returns:
        True if both Nhóm hàng and Thuế VAT equal 'MIX'.
    """
    nhom_hang_idx = col_indices.get("nhom_hang")
    thue_vat_idx = col_indices.get("thue_vat")

    if nhom_hang_idx is None or thue_vat_idx is None:
        return False

    if nhom_hang_idx >= len(row) or thue_vat_idx >= len(row):
        return False

    nhom_hang_val = row[nhom_hang_idx].value
    thue_vat_val = row[thue_vat_idx].value

    nhom_hang_str = str(nhom_hang_val).strip().upper() if nhom_hang_val is not None else ""
    thue_vat_str = str(thue_vat_val).strip().upper() if thue_vat_val is not None else ""

    return nhom_hang_str == "MIX" and thue_vat_str == "MIX"


def count_combo_codes(cell_value) -> tuple[list[str], int]:
    """Count product codes in a multi-line cell value.

    Splits cell value by newline, filters empty lines.

    Args:
        cell_value: The cell value (string or None).

    Returns:
        Tuple of (list of codes, count).
    """
    if cell_value is None:
        return [], 0

    text = str(cell_value)
    lines = [line.strip() for line in text.split("\n") if line.strip()]
    return lines, len(lines)


def _safe_str(value) -> str:
    """Convert a cell value to string, handling None."""
    if value is None:
        return ""
    return str(value).strip()


def classify_rows(filepath: str) -> list[ClassifiedRow]:
    """Classify all rows in an Excel file.

    Opens the workbook, finds the header row, then classifies each data row as
    CATEGORY, PRODUCT, COMBO, COMBO_CHILD, or SKIP.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        List of ClassifiedRow objects.
    """
    # Load workbook twice: data_only=True for computed values, False for font info
    wb_data = openpyxl.load_workbook(filepath, data_only=True)
    wb_font = openpyxl.load_workbook(filepath, data_only=False)

    ws_data = wb_data.active
    ws_font = wb_font.active

    header_row_num = find_header_row(ws_data)
    if header_row_num is None:
        wb_data.close()
        wb_font.close()
        return []

    # Column indices (0-based within row tuple)
    # These are the known positions for this Excel format
    col_indices = {
        "stt": 0,         # Col A
        "ma_hang": 2,     # Col C
        "mo_ta": 3,       # Col D
        "nhom_hang": 6,   # Col G (Nhóm hàng - for MIX detection)
        "thue_vat": 8,    # Col I (Thuế VAT)
        "gia_ban_ra": 10, # Col K (Giá Đại lý đề xuất có VAT)
        "gia_niem_yet": 12, # Col M (Giá bán lẻ đề xuất có VAT)
    }

    data_start_row = header_row_num + 1
    max_col = max(col_indices.values()) + 1

    # Read all data rows from both workbooks
    data_rows = list(ws_data.iter_rows(
        min_row=data_start_row, max_row=ws_data.max_row, min_col=1, max_col=max_col
    ))
    font_rows = list(ws_font.iter_rows(
        min_row=data_start_row, max_row=ws_font.max_row, min_col=1, max_col=max_col
    ))

    results: list[ClassifiedRow] = []
    current_danh_muc = ""
    i = 0

    while i < len(data_rows):
        data_row = data_rows[i]
        font_row = font_rows[i] if i < len(font_rows) else data_row
        excel_row_num = data_start_row + i

        stt_idx = col_indices["stt"]
        stt_val = data_row[stt_idx].value if stt_idx < len(data_row) else None
        # STT must be a number (int/float) or "###" — text values are not valid STT
        has_stt = False
        if stt_val is not None:
            stt_str = str(stt_val).strip()
            if stt_str:
                try:
                    float(stt_str)
                    has_stt = True
                except ValueError:
                    # Could be "###" or "#REF!" from Excel — treat as valid STT
                    if stt_str.startswith("#"):
                        has_stt = True

        # Check for category row: no numeric STT but has text content
        if not has_stt:
            # Find the description/name for this category row
            # Use column D (mo_ta) or the first non-empty cell value
            mo_ta_idx = col_indices["mo_ta"]
            cat_name = ""
            if mo_ta_idx < len(data_row):
                cat_name = _safe_str(data_row[mo_ta_idx].value)
            if not cat_name:
                # Fallback: use first non-empty cell
                for cell in data_row:
                    if cell.value is not None and str(cell.value).strip():
                        cat_name = _safe_str(cell.value)
                        break

            if cat_name:
                current_danh_muc = cat_name
                results.append(ClassifiedRow(
                    row_index=excel_row_num,
                    row_type=RowType.CATEGORY,
                    danh_muc=current_danh_muc,
                    description=cat_name,
                ))
            else:
                results.append(ClassifiedRow(
                    row_index=excel_row_num,
                    row_type=RowType.SKIP,
                    danh_muc=current_danh_muc,
                ))
            i += 1
            continue

        # Check for combo row
        if has_stt and is_combo_row(data_row, col_indices):
            ma_hang_idx = col_indices["ma_hang"]
            mo_ta_idx = col_indices["mo_ta"]
            thue_vat_idx = col_indices["thue_vat"]
            gia_ban_ra_idx = col_indices["gia_ban_ra"]
            gia_niem_yet_idx = col_indices["gia_niem_yet"]

            ma_hang_val = data_row[ma_hang_idx].value if ma_hang_idx < len(data_row) else None
            combo_desc = _safe_str(data_row[mo_ta_idx].value) if mo_ta_idx < len(data_row) else ""
            thue_vat_val = _safe_str(data_row[thue_vat_idx].value) if thue_vat_idx < len(data_row) else ""
            gia_ban_ra_val = _safe_str(data_row[gia_ban_ra_idx].value) if gia_ban_ra_idx < len(data_row) else ""
            gia_niem_yet_val = _safe_str(data_row[gia_niem_yet_idx].value) if gia_niem_yet_idx < len(data_row) else ""

            codes, n_codes = count_combo_codes(ma_hang_val)

            results.append(ClassifiedRow(
                row_index=excel_row_num,
                row_type=RowType.COMBO,
                ma_san_pham=", ".join(codes),
                description=combo_desc,
                thue_vat=thue_vat_val,
                gia_ban_ra=gia_ban_ra_val,
                gia_niem_yet=gia_niem_yet_val,
                danh_muc=current_danh_muc,
                combo_codes=codes,
            ))

            # Consume next N rows as COMBO_CHILD
            for j in range(n_codes):
                child_i = i + 1 + j
                if child_i >= len(data_rows):
                    break

                child_row = data_rows[child_i]
                child_excel_row = data_start_row + child_i

                child_ma = _safe_str(child_row[ma_hang_idx].value) if ma_hang_idx < len(child_row) else ""
                child_desc = _safe_str(child_row[mo_ta_idx].value) if mo_ta_idx < len(child_row) else ""
                child_thue = _safe_str(child_row[thue_vat_idx].value) if thue_vat_idx < len(child_row) else ""
                child_gia_ban = _safe_str(child_row[gia_ban_ra_idx].value) if gia_ban_ra_idx < len(child_row) else ""
                child_gia_niem = _safe_str(child_row[gia_niem_yet_idx].value) if gia_niem_yet_idx < len(child_row) else ""

                # Use child's own description if available, else parent combo description
                effective_desc = child_desc if child_desc else combo_desc

                results.append(ClassifiedRow(
                    row_index=child_excel_row,
                    row_type=RowType.COMBO_CHILD,
                    ma_san_pham=child_ma,
                    description=effective_desc,
                    thue_vat=child_thue,
                    gia_ban_ra=child_gia_ban,
                    gia_niem_yet=child_gia_niem,
                    danh_muc=current_danh_muc,
                    parent_combo_desc=combo_desc,
                ))

            i += 1 + n_codes
            continue

        # Check for product row
        if has_stt:
            ma_hang_idx = col_indices["ma_hang"]
            mo_ta_idx = col_indices["mo_ta"]
            thue_vat_idx = col_indices["thue_vat"]
            gia_ban_ra_idx = col_indices["gia_ban_ra"]
            gia_niem_yet_idx = col_indices["gia_niem_yet"]

            results.append(ClassifiedRow(
                row_index=excel_row_num,
                row_type=RowType.PRODUCT,
                ma_san_pham=_safe_str(data_row[ma_hang_idx].value) if ma_hang_idx < len(data_row) else "",
                description=_safe_str(data_row[mo_ta_idx].value) if mo_ta_idx < len(data_row) else "",
                thue_vat=_safe_str(data_row[thue_vat_idx].value) if thue_vat_idx < len(data_row) else "",
                gia_ban_ra=_safe_str(data_row[gia_ban_ra_idx].value) if gia_ban_ra_idx < len(data_row) else "",
                gia_niem_yet=_safe_str(data_row[gia_niem_yet_idx].value) if gia_niem_yet_idx < len(data_row) else "",
                danh_muc=current_danh_muc,
            ))
            i += 1
            continue

        # SKIP row
        results.append(ClassifiedRow(
            row_index=excel_row_num,
            row_type=RowType.SKIP,
            danh_muc=current_danh_muc,
        ))
        i += 1

    wb_data.close()
    wb_font.close()

    return results


# Standard 13-column output headers
OUTPUT_HEADERS = [
    "ma_san_pham", "ma_phu_kien", "ten_san_pham", "hinh_anh",
    "mo_ta_tinh_nang", "kich_thuoc", "thuong_hieu", "danh_muc",
    "don_vi", "thue_vat", "gia_niem_yet", "gia_mua_vao", "gia_ban_ra",
]


def read_excel_preview(filepath: str) -> dict:
    """Read the first sheet of an Excel file and return ALL rows for preview.

    Returns the complete file content including metadata rows at the top,
    so the preview matches the original Excel file exactly.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        Dict with keys:
            - rows: list of lists of strings (ALL rows from the sheet)
            - row_count: total number of data rows (after header, for info display)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    all_rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        all_rows.append([_safe_str(cell.value) for cell in row])

    # Count data rows (after header) for info
    header_row_num = find_header_row(ws)
    data_row_count = 0
    if header_row_num is not None:
        data_row_count = ws.max_row - header_row_num

    wb.close()
    return {"rows": all_rows, "row_count": data_row_count}


def build_output_rows(
    classified_rows: list[ClassifiedRow],
    llm_results: dict[int, dict],
) -> list[dict]:
    """Merge classified Excel data with LLM results into 13-column output rows.

    Args:
        classified_rows: List of ClassifiedRow from classify_rows().
        llm_results: Dict mapping row_index → {ten_san_pham, mo_ta_tinh_nang,
                      kich_thuoc, thuong_hieu}.

    Returns:
        List of dicts, each with exactly 13 string-valued keys matching OUTPUT_HEADERS.
        CATEGORY and SKIP rows are excluded from output.
    """
    output: list[dict] = []
    stt_counter = 0

    for row in classified_rows:
        if row.row_type in (RowType.CATEGORY, RowType.SKIP):
            continue

        llm = llm_results.get(row.row_index, {})
        ten_san_pham = str(llm.get("ten_san_pham", ""))
        mo_ta_tinh_nang = str(llm.get("mo_ta_tinh_nang", ""))
        kich_thuoc = str(llm.get("kich_thuoc", ""))
        thuong_hieu = str(llm.get("thuong_hieu", ""))

        if row.row_type == RowType.PRODUCT:
            stt_counter += 1
            output.append({
                "_stt": str(stt_counter),
                "ma_san_pham": str(row.ma_san_pham),
                "ma_phu_kien": "",
                "ten_san_pham": ten_san_pham,
                "hinh_anh": "",
                "mo_ta_tinh_nang": mo_ta_tinh_nang,
                "kich_thuoc": kich_thuoc,
                "thuong_hieu": thuong_hieu,
                "danh_muc": str(row.danh_muc),
                "don_vi": "",
                "thue_vat": str(row.thue_vat),
                "gia_niem_yet": str(row.gia_niem_yet),
                "gia_mua_vao": "",
                "gia_ban_ra": str(row.gia_ban_ra),
            })

        elif row.row_type == RowType.COMBO:
            stt_counter += 1
            # Combo summary row: all codes joined, combo totals, other fields empty
            output.append({
                "_stt": str(stt_counter),
                "ma_san_pham": str(row.ma_san_pham),
                "ma_phu_kien": "",
                "ten_san_pham": "",
                "hinh_anh": "",
                "mo_ta_tinh_nang": "",
                "kich_thuoc": "",
                "thuong_hieu": "",
                "danh_muc": "",
                "don_vi": "",
                "thue_vat": "",
                "gia_niem_yet": str(row.gia_niem_yet),
                "gia_mua_vao": "",
                "gia_ban_ra": str(row.gia_ban_ra),
            })

        elif row.row_type == RowType.COMBO_CHILD:
            # Same STT as parent combo (stt_counter not incremented)
            output.append({
                "_stt": str(stt_counter),
                "ma_san_pham": "",
                "ma_phu_kien": str(row.ma_san_pham),
                "ten_san_pham": ten_san_pham,
                "hinh_anh": "",
                "mo_ta_tinh_nang": mo_ta_tinh_nang,
                "kich_thuoc": kich_thuoc,
                "thuong_hieu": thuong_hieu,
                "danh_muc": str(row.danh_muc),
                "don_vi": "",
                "thue_vat": str(row.thue_vat),
                "gia_niem_yet": str(row.gia_niem_yet),
                "gia_mua_vao": "",
                "gia_ban_ra": str(row.gia_ban_ra),
            })

    return output


def build_descriptions_batch(
    classified_rows: list[ClassifiedRow],
) -> list[tuple[int, str]]:
    """Build list of (row_index, description) tuples for rows needing LLM analysis.

    Filters to PRODUCT and COMBO_CHILD rows only (COMBO summary, CATEGORY, and
    SKIP rows are excluded). For COMBO_CHILD, uses the effective description
    (own description if available, otherwise parent_combo_desc). Rows with empty
    descriptions are skipped.

    Args:
        classified_rows: List of ClassifiedRow from classify_rows().

    Returns:
        List of (row_index, description) tuples for non-empty descriptions.
    """
    result: list[tuple[int, str]] = []

    for row in classified_rows:
        if row.row_type not in (RowType.PRODUCT, RowType.COMBO_CHILD):
            continue

        # For COMBO_CHILD, use effective description (own or parent)
        if row.row_type == RowType.COMBO_CHILD:
            desc = row.description if row.description else row.parent_combo_desc
        else:
            desc = row.description

        # Skip rows with empty description
        if not desc or not desc.strip():
            continue

        result.append((row.row_index, desc))

    return result


def call_llm_text_batch(
    descriptions: list[str], model: str, api_key: str
) -> tuple[list[dict], dict]:
    """Send a batch of text descriptions to OpenRouter API for parsing.

    Each description is parsed into: ten_san_pham, mo_ta_tinh_nang,
    kich_thuoc, thuong_hieu. No images are sent.

    Args:
        descriptions: List of product description strings.
        model: Model name for OpenRouter API.
        api_key: OpenRouter API key.

    Returns:
        Tuple of (list of dicts with 4 fields, usage dict).
        On error, returns list of empty dicts and empty usage.
    """
    empty_row = {"ten_san_pham": "", "mo_ta_tinh_nang": "", "kich_thuoc": "", "thuong_hieu": ""}
    empty_results = [dict(empty_row) for _ in descriptions]

    # If all descriptions are empty, return without calling API
    if not descriptions or all(not d.strip() for d in descriptions):
        return empty_results, {}

    prompt = f"""Bạn là chuyên gia phân tích mô tả sản phẩm.
Cho danh sách mô tả sản phẩm dưới đây, hãy trích xuất cho MỖI mô tả:
- "ten_san_pham": Tên sản phẩm ngắn gọn
- "mo_ta_tinh_nang": Mô tả tính năng chi tiết
- "kich_thuoc": Kích thước (nếu có)
- "thuong_hieu": Thương hiệu (nếu có)

Trả về JSON array với đúng {len(descriptions)} phần tử, theo thứ tự tương ứng.
Nếu mô tả rỗng hoặc không có thông tin cho một trường, trả về chuỗi rỗng "" cho trường đó.
CHỈ trả về JSON array, không giải thích gì thêm.

Danh sách mô tả:
{json.dumps(descriptions, ensure_ascii=False)}"""

    try:
        resp = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
            json={
                "model": model,
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 16000,
                "temperature": 0.1,
            },
            timeout=300,
        )

        if resp.status_code != 200:
            log.error(f"LLM API error {resp.status_code}: {resp.text[:500]}")
            return empty_results, {}

        data = resp.json()
        if "error" in data:
            log.error(f"LLM API returned error: {data['error']}")
            return empty_results, {}

        content_text = data["choices"][0]["message"]["content"]
        usage = data.get("usage", data["choices"][0].get("usage", {}))

        # Parse JSON from response (strip markdown fences if present)
        cleaned = content_text.strip()
        if cleaned.startswith("```"):
            cleaned = cleaned.split("\n", 1)[1]
        if cleaned.endswith("```"):
            cleaned = cleaned.rsplit("```", 1)[0]
        cleaned = cleaned.strip()

        parsed = json.loads(cleaned)
        if not isinstance(parsed, list):
            log.warning("LLM response is not a JSON array")
            return empty_results, usage

        # Validate count matches
        if len(parsed) != len(descriptions):
            log.warning(
                f"LLM returned {len(parsed)} items, expected {len(descriptions)}"
            )
            return empty_results, usage

        # Normalize each result to ensure all 4 fields exist as strings
        results = []
        for item in parsed:
            if not isinstance(item, dict):
                results.append(dict(empty_row))
                continue
            results.append({
                "ten_san_pham": str(item.get("ten_san_pham", "")),
                "mo_ta_tinh_nang": str(item.get("mo_ta_tinh_nang", "")),
                "kich_thuoc": str(item.get("kich_thuoc", "")),
                "thuong_hieu": str(item.get("thuong_hieu", "")),
            })

        return results, usage

    except requests.Timeout:
        log.error("LLM API request timed out")
        return empty_results, {}
    except json.JSONDecodeError as e:
        log.error(f"Failed to parse LLM response as JSON: {e}")
        return empty_results, {}
    except Exception as e:
        log.error(f"Unexpected error calling LLM: {e}")
        return empty_results, {}
