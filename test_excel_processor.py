"""Unit tests for excel_processor.py — tasks 2.1, 2.2, 2.3, 4.1."""

import json
import os
import tempfile

import openpyxl
import pytest

from excel_processor import (
    ClassifiedRow,
    OUTPUT_HEADERS,
    RowType,
    build_descriptions_batch,
    build_output_rows,
    read_excel_preview,
)


# ── Helpers ──────────────────────────────────────────────────────────────────


def _create_xlsx(rows: list[list], bold_rows: set[int] | None = None) -> str:
    """Create a temp .xlsx file with given rows. bold_rows is a set of 0-based row indices."""
    bold_rows = bold_rows or set()
    wb = openpyxl.Workbook()
    ws = wb.active
    from openpyxl.styles import Font

    for r_idx, row_data in enumerate(rows):
        for c_idx, val in enumerate(row_data):
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=val)
            if r_idx in bold_rows:
                cell.font = Font(bold=True)

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    wb.close()
    return path


# ── Tests for read_excel_preview (Task 2.1) ─────────────────────────────────


class TestReadExcelPreview:
    def test_basic_preview(self):
        """Headers + data rows returned correctly, all values as strings."""
        rows = [
            ["STT", "Nhóm hàng", "Mã hàng", "Mô tả"],
            [1, "A", "CODE1", "Product 1"],
            [2, "B", "CODE2", "Product 2"],
        ]
        path = _create_xlsx(rows)
        try:
            result = read_excel_preview(path)
            assert result["headers"] == ["STT", "Nhóm hàng", "Mã hàng", "Mô tả"]
            assert result["row_count"] == 2
            assert len(result["rows"]) == 2
            assert result["rows"][0] == ["1", "A", "CODE1", "Product 1"]
            assert result["rows"][1] == ["2", "B", "CODE2", "Product 2"]
        finally:
            os.unlink(path)

    def test_all_values_are_strings(self):
        """Numeric and None values are converted to strings."""
        rows = [
            ["STT", "Price"],
            [1, 12345.67],
            [None, None],
        ]
        path = _create_xlsx(rows)
        try:
            result = read_excel_preview(path)
            for row in result["rows"]:
                for val in row:
                    assert isinstance(val, str)
        finally:
            os.unlink(path)

    def test_no_header_row(self):
        """When no STT header exists, returns all rows with empty headers."""
        rows = [
            ["A", "B", "C"],
            [1, 2, 3],
        ]
        path = _create_xlsx(rows)
        try:
            result = read_excel_preview(path)
            assert result["headers"] == []
            assert result["row_count"] == 2
        finally:
            os.unlink(path)

    def test_header_not_on_first_row(self):
        """Header row can be on any row, not just row 1."""
        rows = [
            ["Metadata", "info"],
            ["Company", "XYZ"],
            ["STT", "Mã hàng", "Mô tả"],
            [1, "CODE1", "Desc 1"],
        ]
        path = _create_xlsx(rows)
        try:
            result = read_excel_preview(path)
            assert result["headers"] == ["STT", "Mã hàng", "Mô tả"]
            assert result["row_count"] == 1
            assert result["rows"][0] == ["1", "CODE1", "Desc 1"]
        finally:
            os.unlink(path)

    def test_empty_sheet(self):
        """Empty sheet (no STT header) returns empty headers."""
        wb = openpyxl.Workbook()
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        wb.save(path)
        wb.close()
        try:
            result = read_excel_preview(path)
            assert result["headers"] == []
            # openpyxl reports max_row=1 for a new workbook, so 1 row is returned
            assert isinstance(result["row_count"], int)
        finally:
            os.unlink(path)


# ── Tests for build_output_rows (Task 2.2) ──────────────────────────────────


class TestBuildOutputRows:
    def test_product_row(self):
        """PRODUCT row fills all 13 fields correctly."""
        rows = [
            ClassifiedRow(
                row_index=5, row_type=RowType.PRODUCT,
                ma_san_pham="ABC123", description="Some desc",
                thue_vat="10%", gia_ban_ra="500000", gia_niem_yet="600000",
                danh_muc="Vòi sen",
            ),
        ]
        llm = {5: {"ten_san_pham": "Vòi ABC", "mo_ta_tinh_nang": "Chống gỉ", "kich_thuoc": "30cm", "thuong_hieu": "INAX"}}
        result = build_output_rows(rows, llm)

        assert len(result) == 1
        r = result[0]
        assert r["ma_san_pham"] == "ABC123"
        assert r["ma_phu_kien"] == ""
        assert r["ten_san_pham"] == "Vòi ABC"
        assert r["hinh_anh"] == ""
        assert r["mo_ta_tinh_nang"] == "Chống gỉ"
        assert r["kich_thuoc"] == "30cm"
        assert r["thuong_hieu"] == "INAX"
        assert r["danh_muc"] == "Vòi sen"
        assert r["don_vi"] == ""
        assert r["thue_vat"] == "10%"
        assert r["gia_niem_yet"] == "600000"
        assert r["gia_mua_vao"] == ""
        assert r["gia_ban_ra"] == "500000"

    def test_combo_and_children(self):
        """COMBO row has all codes + totals; COMBO_CHILD has ma_phu_kien."""
        rows = [
            ClassifiedRow(
                row_index=10, row_type=RowType.COMBO,
                ma_san_pham="C1, C2", thue_vat="MIX",
                gia_ban_ra="1000000", gia_niem_yet="1200000",
                danh_muc="Combo", combo_codes=["C1", "C2"],
            ),
            ClassifiedRow(
                row_index=11, row_type=RowType.COMBO_CHILD,
                ma_san_pham="C1", description="Child 1 desc",
                thue_vat="10%", gia_ban_ra="400000", gia_niem_yet="500000",
                danh_muc="Combo", parent_combo_desc="Combo desc",
            ),
            ClassifiedRow(
                row_index=12, row_type=RowType.COMBO_CHILD,
                ma_san_pham="C2", description="Child 2 desc",
                thue_vat="10%", gia_ban_ra="600000", gia_niem_yet="700000",
                danh_muc="Combo", parent_combo_desc="Combo desc",
            ),
        ]
        llm = {
            11: {"ten_san_pham": "Child1", "mo_ta_tinh_nang": "F1", "kich_thuoc": "S1", "thuong_hieu": "B1"},
            12: {"ten_san_pham": "Child2", "mo_ta_tinh_nang": "F2", "kich_thuoc": "S2", "thuong_hieu": "B2"},
        }
        result = build_output_rows(rows, llm)

        assert len(result) == 3

        # Combo summary
        combo = result[0]
        assert combo["ma_san_pham"] == "C1, C2"
        assert combo["ma_phu_kien"] == ""
        assert combo["gia_niem_yet"] == "1200000"
        assert combo["gia_ban_ra"] == "1000000"
        assert combo["ten_san_pham"] == ""

        # Child 1
        c1 = result[1]
        assert c1["ma_san_pham"] == ""
        assert c1["ma_phu_kien"] == "C1"
        assert c1["ten_san_pham"] == "Child1"
        assert c1["thue_vat"] == "10%"

        # Child 2
        c2 = result[2]
        assert c2["ma_phu_kien"] == "C2"

    def test_category_and_skip_excluded(self):
        """CATEGORY and SKIP rows are not in output."""
        rows = [
            ClassifiedRow(row_index=1, row_type=RowType.CATEGORY, danh_muc="Cat"),
            ClassifiedRow(row_index=2, row_type=RowType.SKIP),
            ClassifiedRow(
                row_index=3, row_type=RowType.PRODUCT,
                ma_san_pham="X", thue_vat="5%",
                gia_ban_ra="100", gia_niem_yet="200", danh_muc="Cat",
            ),
        ]
        result = build_output_rows(rows, {})
        assert len(result) == 1
        assert result[0]["ma_san_pham"] == "X"

    def test_all_values_are_strings(self):
        """Every value in every output dict must be a string."""
        rows = [
            ClassifiedRow(
                row_index=5, row_type=RowType.PRODUCT,
                ma_san_pham="A", thue_vat="10",
                gia_ban_ra="100", gia_niem_yet="200", danh_muc="D",
            ),
        ]
        result = build_output_rows(rows, {})
        for r in result:
            for k, v in r.items():
                assert isinstance(v, str), f"Key {k} has non-string value: {type(v)}"

    def test_output_has_13_keys(self):
        """Each output dict has exactly 13 keys matching OUTPUT_HEADERS."""
        rows = [
            ClassifiedRow(
                row_index=5, row_type=RowType.PRODUCT,
                ma_san_pham="A", thue_vat="10",
                gia_ban_ra="100", gia_niem_yet="200", danh_muc="D",
            ),
        ]
        result = build_output_rows(rows, {})
        assert len(result[0]) == 13
        assert set(result[0].keys()) == set(OUTPUT_HEADERS)

    def test_always_empty_fields(self):
        """hinh_anh, don_vi, gia_mua_vao are always empty strings."""
        rows = [
            ClassifiedRow(
                row_index=5, row_type=RowType.PRODUCT,
                ma_san_pham="A", thue_vat="10",
                gia_ban_ra="100", gia_niem_yet="200", danh_muc="D",
            ),
        ]
        result = build_output_rows(rows, {})
        assert result[0]["hinh_anh"] == ""
        assert result[0]["don_vi"] == ""
        assert result[0]["gia_mua_vao"] == ""

    def test_missing_llm_results_default_empty(self):
        """When llm_results has no entry for a row, LLM fields default to empty."""
        rows = [
            ClassifiedRow(
                row_index=5, row_type=RowType.PRODUCT,
                ma_san_pham="A", thue_vat="10",
                gia_ban_ra="100", gia_niem_yet="200", danh_muc="D",
            ),
        ]
        result = build_output_rows(rows, {})
        assert result[0]["ten_san_pham"] == ""
        assert result[0]["mo_ta_tinh_nang"] == ""
        assert result[0]["kich_thuoc"] == ""
        assert result[0]["thuong_hieu"] == ""


# ── Tests for build_descriptions_batch (Task 2.3) ───────────────────────────


class TestBuildDescriptionsBatch:
    def test_product_rows_included(self):
        """PRODUCT rows with descriptions are included."""
        rows = [
            ClassifiedRow(row_index=5, row_type=RowType.PRODUCT, description="Desc A"),
            ClassifiedRow(row_index=6, row_type=RowType.PRODUCT, description="Desc B"),
        ]
        result = build_descriptions_batch(rows)
        assert result == [(5, "Desc A"), (6, "Desc B")]

    def test_combo_excluded(self):
        """COMBO rows are NOT included (only COMBO_CHILD)."""
        rows = [
            ClassifiedRow(row_index=10, row_type=RowType.COMBO, description="Combo desc"),
        ]
        result = build_descriptions_batch(rows)
        assert result == []

    def test_category_and_skip_excluded(self):
        """CATEGORY and SKIP rows are excluded."""
        rows = [
            ClassifiedRow(row_index=1, row_type=RowType.CATEGORY, description="Cat"),
            ClassifiedRow(row_index=2, row_type=RowType.SKIP, description="Skip"),
        ]
        result = build_descriptions_batch(rows)
        assert result == []

    def test_combo_child_own_description(self):
        """COMBO_CHILD with own description uses it."""
        rows = [
            ClassifiedRow(
                row_index=11, row_type=RowType.COMBO_CHILD,
                description="Own desc", parent_combo_desc="Parent desc",
            ),
        ]
        result = build_descriptions_batch(rows)
        assert result == [(11, "Own desc")]

    def test_combo_child_falls_back_to_parent(self):
        """COMBO_CHILD with empty description uses parent_combo_desc."""
        rows = [
            ClassifiedRow(
                row_index=11, row_type=RowType.COMBO_CHILD,
                description="", parent_combo_desc="Parent desc",
            ),
        ]
        result = build_descriptions_batch(rows)
        assert result == [(11, "Parent desc")]

    def test_empty_description_skipped(self):
        """Rows with empty description (and no parent fallback) are skipped."""
        rows = [
            ClassifiedRow(row_index=5, row_type=RowType.PRODUCT, description=""),
            ClassifiedRow(row_index=6, row_type=RowType.PRODUCT, description="  "),
            ClassifiedRow(
                row_index=7, row_type=RowType.COMBO_CHILD,
                description="", parent_combo_desc="",
            ),
        ]
        result = build_descriptions_batch(rows)
        assert result == []

    def test_mixed_rows(self):
        """Full scenario with mixed row types."""
        rows = [
            ClassifiedRow(row_index=1, row_type=RowType.CATEGORY, description="Cat"),
            ClassifiedRow(row_index=2, row_type=RowType.PRODUCT, description="Prod A"),
            ClassifiedRow(row_index=3, row_type=RowType.COMBO, description="Combo X"),
            ClassifiedRow(
                row_index=4, row_type=RowType.COMBO_CHILD,
                description="Child own", parent_combo_desc="Combo X",
            ),
            ClassifiedRow(
                row_index=5, row_type=RowType.COMBO_CHILD,
                description="", parent_combo_desc="Combo X",
            ),
            ClassifiedRow(row_index=6, row_type=RowType.SKIP),
            ClassifiedRow(row_index=7, row_type=RowType.PRODUCT, description=""),
        ]
        result = build_descriptions_batch(rows)
        assert result == [
            (2, "Prod A"),
            (4, "Child own"),
            (5, "Combo X"),
        ]


# ── Tests for call_llm_text_batch (Task 4.1) ────────────────────────────────

from unittest.mock import patch, MagicMock
from excel_processor import call_llm_text_batch


class TestCallLlmTextBatch:
    def test_all_empty_descriptions_no_api_call(self):
        """When all descriptions are empty, returns empty dicts without calling API."""
        results, usage = call_llm_text_batch(["", "  ", ""], "model", "key")
        assert len(results) == 3
        assert usage == {}
        for r in results:
            assert r == {"ten_san_pham": "", "mo_ta_tinh_nang": "", "kich_thuoc": "", "thuong_hieu": ""}

    def test_empty_list_no_api_call(self):
        """Empty list returns empty results without calling API."""
        results, usage = call_llm_text_batch([], "model", "key")
        assert results == []
        assert usage == {}

    @patch("excel_processor.requests.post")
    def test_successful_response(self, mock_post):
        """Successful API call returns parsed results and usage."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {
            "choices": [{"message": {"content": json.dumps([
                {"ten_san_pham": "Vòi sen", "mo_ta_tinh_nang": "Chống gỉ", "kich_thuoc": "30cm", "thuong_hieu": "INAX"},
                {"ten_san_pham": "Lavabo", "mo_ta_tinh_nang": "Cao cấp", "kich_thuoc": "50x40", "thuong_hieu": "TOTO"},
            ], ensure_ascii=False)}}],
            "usage": {"prompt_tokens": 100, "completion_tokens": 50},
        }
        mock_post.return_value = mock_resp

        results, usage = call_llm_text_batch(["Desc 1", "Desc 2"], "model", "key")
        assert len(results) == 2
        assert results[0]["ten_san_pham"] == "Vòi sen"
        assert results[1]["thuong_hieu"] == "TOTO"
        assert usage["prompt_tokens"] == 100

    @patch("excel_processor.requests.post")
    def test_api_error_status(self, mock_post):
        """Non-200 status returns empty dicts."""
        mock_resp = MagicMock()
        mock_resp.status_code = 500
        mock_resp.text = "Internal Server Error"
        mock_post.return_value = mock_resp

        results, usage = call_llm_text_batch(["Desc"], "model", "key")
        assert len(results) == 1
        assert results[0]["ten_san_pham"] == ""
        assert usage == {}

    @patch("excel_processor.requests.post")
    def test_timeout_returns_empty(self, mock_post):
        """Timeout returns empty dicts."""
        import requests as req
        mock_post.side_effect = req.Timeout("timed out")

        results, usage = call_llm_text_batch(["Desc"], "model", "key")
        assert len(results) == 1
        assert results[0]["ten_san_pham"] == ""
        assert usage == {}

    @patch("excel_processor.requests.post")
    def test_invalid_json_response(self, mock_post):
        """Invalid JSON in LLM response returns empty dicts."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {
            "choices": [{"message": {"content": "not valid json at all"}}],
            "usage": {"prompt_tokens": 10, "completion_tokens": 5},
        }
        mock_post.return_value = mock_resp

        results, usage = call_llm_text_batch(["Desc"], "model", "key")
        assert len(results) == 1
        assert results[0]["ten_san_pham"] == ""

    @patch("excel_processor.requests.post")
    def test_wrong_count_returns_empty(self, mock_post):
        """When LLM returns wrong number of items, returns empty dicts."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {
            "choices": [{"message": {"content": json.dumps([
                {"ten_san_pham": "A", "mo_ta_tinh_nang": "", "kich_thuoc": "", "thuong_hieu": ""},
            ])}}],
            "usage": {},
        }
        mock_post.return_value = mock_resp

        results, usage = call_llm_text_batch(["Desc 1", "Desc 2"], "model", "key")
        assert len(results) == 2
        assert results[0]["ten_san_pham"] == ""
        assert results[1]["ten_san_pham"] == ""

    @patch("excel_processor.requests.post")
    def test_markdown_fenced_json(self, mock_post):
        """LLM response wrapped in markdown code fences is handled."""
        json_content = json.dumps([
            {"ten_san_pham": "Test", "mo_ta_tinh_nang": "F", "kich_thuoc": "S", "thuong_hieu": "B"},
        ], ensure_ascii=False)
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {
            "choices": [{"message": {"content": f"```json\n{json_content}\n```"}}],
            "usage": {},
        }
        mock_post.return_value = mock_resp

        results, usage = call_llm_text_batch(["Desc"], "model", "key")
        assert results[0]["ten_san_pham"] == "Test"

    @patch("excel_processor.requests.post")
    def test_api_error_field_in_response(self, mock_post):
        """API response with 'error' field returns empty dicts."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {"error": {"message": "rate limited"}}
        mock_post.return_value = mock_resp

        results, usage = call_llm_text_batch(["Desc"], "model", "key")
        assert len(results) == 1
        assert results[0]["ten_san_pham"] == ""
        assert usage == {}

    @patch("excel_processor.requests.post")
    def test_sends_text_only_no_images(self, mock_post):
        """Verify the API call sends text-only content, no image_url."""
        mock_resp = MagicMock()
        mock_resp.status_code = 200
        mock_resp.json.return_value = {
            "choices": [{"message": {"content": json.dumps([
                {"ten_san_pham": "A", "mo_ta_tinh_nang": "", "kich_thuoc": "", "thuong_hieu": ""},
            ])}}],
            "usage": {},
        }
        mock_post.return_value = mock_resp

        call_llm_text_batch(["Desc"], "test-model", "test-key")

        call_args = mock_post.call_args
        payload = call_args.kwargs.get("json") or call_args[1].get("json")
        messages = payload["messages"]
        # Content should be a plain string (text prompt), not a list with image_url
        assert isinstance(messages[0]["content"], str)
        assert "image_url" not in messages[0]["content"]
